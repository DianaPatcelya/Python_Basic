from openpyxl import load_workbook, Workbook
from datetime import datetime
from person import Person


class Database:
    def __init__(self):
        self.people = []

    def add_person(self, person):
        self.people.append(person)

    def search_people(self, query):
        query = query.lower()
        results = []
        for person in self.people:
            if person.matches_query(query):
                results.append(person)
        return results

    def save_to_excel(self, filename):
        wb = Workbook()
        ws = wb.active
        for person in self.people:
            ws.append([person.first_name, person.surname, person.paternal_name, person.date_birth.strftime("%d.%m.%Y"),
                       person.date_death.strftime("%d.%m.%Y") if person.date_death else '', person.gender])
        wb.save(filename)

    def load_from_excel(self, filename):
        workbook = load_workbook(filename)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            first_name, surname, paternal_name, date_birth, date_death, gender = row
            if date_birth:
                date_birth = datetime.strptime(date_birth, "%d.%m.%Y")
            if date_death:
                date_death = datetime.strptime(date_death, "%d.%m.%Y")
            person = Person(first_name, gender, date_birth, surname, paternal_name, date_death)
            self.people.append(person)

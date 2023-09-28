import csv
import openpyxl

work_book = openpyxl.Workbook()
default_sheet = work_book["Sheet"]
work_book.remove(default_sheet)
sheet = work_book.create_sheet("ДЗ20_Excel")

with open("ДЗ_19.csv", mode='r', encoding='utf-8') as f:
    reader = csv.reader(f, delimiter=',')

    for row_index, row in enumerate(reader, start=1):
        sheet.cell(row=row_index + 1, column=1, value=f"Person {row_index}")

        del row[2]

        for col_index, cell_value in enumerate(row, start=2):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.value = cell_value

sheet.delete_rows(row_index + 1)

print("Завершено")
work_book.save('ДЗ_20.xlsx')

